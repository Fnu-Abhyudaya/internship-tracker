"""Main entry point for the internship scraper."""

import os
import sys
import asyncio
import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from .scrapers.base_scraper import JobPosting
from .scrapers.company_configs import get_all_scrapers
from .scrapers.playwright_scraper import shutdown_browser
from .email_sender import send_email, send_no_results_email
from .utils.keywords import matches_role_keywords, is_us_location
from .utils.seen_tracker import (
    load_seen, prune_old, filter_new_postings,
    mark_as_seen, save_seen
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('scraper.log', mode='w'),
    ]
)
logger = logging.getLogger(__name__)

RECIPIENT_EMAIL = os.environ.get(
    'RECIPIENT_EMAIL', 'abhyudaya.manipal@gmail.com'
)
OUTPUT_DIR = Path('output')
MAX_CONCURRENT = 5


async def run_scraper_with_semaphore(scraper, semaphore, stats):
    async with semaphore:
        try:
            results = await asyncio.wait_for(
                scraper.run(), timeout=180
            )
            stats[scraper.company_name] = len(results)
            return results
        except asyncio.TimeoutError:
            logger.warning(f"[{scraper.company_name}] Timed out")
            stats[scraper.company_name] = 'TIMEOUT'
            return []
        except Exception as e:
            logger.error(f"[{scraper.company_name}] Error: {e}")
            stats[scraper.company_name] = f'ERROR: {e}'
            return []


async def run_all_scrapers():
    try:
        scrapers = get_all_scrapers()
    except Exception as e:
        logger.error(f"Failed to load scrapers: {e}")
        logger.error(traceback.format_exc())
        return [], {}

    logger.info(f"Starting {len(scrapers)} scrapers...")
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    stats = {}
    tasks = [
        run_scraper_with_semaphore(s, semaphore, stats)
        for s in scrapers
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    all_postings = []
    for i, result in enumerate(results):
        if isinstance(result, list):
            all_postings.extend(result)
        elif isinstance(result, Exception):
            logger.error(
                f"Scraper {scrapers[i].company_name} "
                f"failed: {result}"
            )

    try:
        await shutdown_browser()
    except Exception as e:
        logger.warning(f"Browser shutdown error: {e}")

    logger.info(f"Total raw postings: {len(all_postings)}")
    return all_postings, stats


def final_filter(postings):
    """Filter by role keywords and US location."""
    filtered = []
    rejected_kw = 0
    rejected_loc = 0
    for p in postings:
        if not matches_role_keywords(p.title):
            rejected_kw += 1
            continue
        if p.location and p.location != 'N/A':
            if not is_us_location(p.location):
                rejected_loc += 1
                continue
        filtered.append(p)
    logger.info(
        f"Filter results: kept {len(filtered)}, "
        f"rejected {rejected_kw} (keyword), "
        f"rejected {rejected_loc} (location)"
    )
    return filtered


def deduplicate(postings):
    """Remove duplicates within this run."""
    seen = set()
    unique = []
    for p in postings:
        key = (
            p.title.lower().strip(),
            p.company.lower().strip(),
            p.url.lower().strip(),
        )
        if key not in seen:
            seen.add(key)
            unique.append(p)
    logger.info(
        f"Within-run dedup: {len(postings)} -> {len(unique)}"
    )
    return unique


def log_company_stats(stats):
    """Print a per-company summary to the log."""
    logger.info("=" * 60)
    logger.info("PER-COMPANY RESULTS:")
    logger.info("=" * 60)
    for company in sorted(stats.keys()):
        count = stats[company]
        if isinstance(count, int):
            marker = "OK " if count > 0 else "-- "
            logger.info(f"  {marker} {company}: {count}")
        else:
            logger.info(f"  ERR {company}: {count}")
    logger.info("=" * 60)


def sort_newest_first(postings):
    """Sort postings by date_posted DESC (newest first).
    Postings with no date go to the bottom."""
    def sort_key(p):
        date = p.date_posted or ''
        # Empty dates sort to end; non-empty sort by reverse string
        return (0 if date else 1, date)

    # Sort with non-empty dates first, then by date desc
    with_date = [p for p in postings if p.date_posted]
    without_date = [p for p in postings if not p.date_posted]

    with_date.sort(
        key=lambda p: p.date_posted or '',
        reverse=True
    )
    without_date.sort(
        key=lambda p: (p.company.lower(), p.title.lower())
    )

    return with_date + without_date


def create_excel(postings):
    """Build and save the styled Excel report."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime('%Y-%m-%d')
    filepath = OUTPUT_DIR / f'internship_postings_{today}.xlsx'

    # Sort newest-first before writing
    postings = sort_newest_first(postings)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Internship Postings'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(
        start_color='2563EB', end_color='2563EB',
        fill_type='solid'
    )
    header_align = Alignment(
        horizontal='center', vertical='center'
    )
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['S.No', 'Role', 'Company', 'Link',
               'Date Posted', 'Location']
    widths = [8, 55, 25, 60, 18, 30]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{max(len(postings) + 1, 2)}'

    alt_fill = PatternFill(
        start_color='EFF6FF', end_color='EFF6FF',
        fill_type='solid'
    )
    link_font = Font(
        color='2563EB', underline='single', size=11
    )
    data_align = Alignment(vertical='center', wrap_text=True)

    for row_num, posting in enumerate(postings, 2):
        row_data = [
            row_num - 1,
            posting.title,
            posting.company,
            posting.url,
            posting.date_posted or 'N/A',
            posting.location or 'N/A',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = data_align
            cell.border = thin
            if row_num % 2 == 0:
                cell.fill = alt_fill
            if col == 4 and value and value != 'N/A':
                cell.font = link_font
                cell.hyperlink = value

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Internship Tracker Summary'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'] = f'Date: {today}'
    ws2['A4'] = f'Total Postings: {len(postings)}'
    companies = set(p.company for p in postings)
    ws2['A5'] = f'Companies: {len(companies)}'
    ws2['A6'] = f'Generated: {datetime.now().isoformat()}'
    ws2['A7'] = 'Sorted by: Most recent first'
    ws2['A9'] = 'By Company:'
    ws2['A9'].font = Font(bold=True)
    ws2['A10'] = 'Company'
    ws2['B10'] = 'Count'
    ws2['A10'].font = Font(bold=True)
    ws2['B10'].font = Font(bold=True)

    counts = {}
    for p in postings:
        counts[p.company] = counts.get(p.company, 0) + 1
    for i, (co, cnt) in enumerate(sorted(counts.items()), 11):
        ws2[f'A{i}'] = co
        ws2[f'B{i}'] = cnt

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 12

    wb.save(filepath)
    logger.info(f"Saved: {filepath}")
    return str(filepath)


def main():
    logger.info("=" * 50)
    logger.info("INTERNSHIP TRACKER STARTING")
    logger.info(f"Time: {datetime.now().isoformat()}")
    logger.info("=" * 50)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load previously-seen jobs and prune old entries
    seen = load_seen()
    seen = prune_old(seen)

    try:
        all_postings, stats = asyncio.run(run_all_scrapers())
    except Exception as e:
        logger.error(f"Scraping crashed: {e}")
        logger.error(traceback.format_exc())
        all_postings, stats = [], {}

    log_company_stats(stats)

    # Apply role + location filter
    filtered = final_filter(all_postings)
    # Dedupe within this run
    unique = deduplicate(filtered)
    # Filter out anything already seen in previous runs
    new_jobs = filter_new_postings(unique, seen)

    # Mark ALL of today's matching jobs as seen
    # (so they won't be re-emailed even if no new ones today)
    seen = mark_as_seen(unique, seen)
    save_seen(seen)

    if not new_jobs:
        logger.info("No NEW postings since last run.")
        try:
            create_excel([])
        except Exception as e:
            logger.error(f"Excel error: {e}")
        try:
            send_no_results_email(RECIPIENT_EMAIL)
        except Exception as e:
            logger.error(f"Email error: {e}")
        return

    try:
        filepath = create_excel(new_jobs)
    except Exception as e:
        logger.error(f"Failed to create Excel: {e}")
        return

    companies = set(p.company for p in new_jobs)
    try:
        send_email(
            filepath=filepath,
            recipient=RECIPIENT_EMAIL,
            total_jobs=len(new_jobs),
            total_companies=len(companies),
        )
        logger.info(f"Report sent with {len(new_jobs)} NEW jobs!")
    except Exception as e:
        logger.error(f"Email failed: {e}")
        logger.error(traceback.format_exc())

    logger.info("DONE")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logger.error(f"FATAL: {e}")
        logger.error(traceback.format_exc())
        sys.exit(0)
